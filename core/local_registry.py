"""
Local GTP Registry — persists all processed cables in output/gtp_registry.xlsx.

Schema (one row per cable):
  GTP No. | Item No. | Description | Config | Conductor | Cores | Area (mm²) |
  Voltage | Margin % | Price A (Rs/km) | Price B (Rs/km) | Price C (Rs/km) |
  Processed At

The Margin % column is user-editable (highlighted yellow). The agent reads it
before pricing so changes take effect on the next run.
"""

import os
from datetime import datetime
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("openpyxl not installed — run: pip install openpyxl")

REGISTRY_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "output", "gtp_registry.xlsx"
)

DEFAULT_MARGIN = 20.0

HEADERS = [
    "GTP No.", "Item No.", "Description", "Config",
    "Conductor", "Cores", "Area (mm²)", "Voltage (kV)",
    "Margin %",
    "Price A (Rs/km)", "Price B (Rs/km)", "Price C (Rs/km)",
    "Processed At",
]

COL_MARGIN    = HEADERS.index("Margin %") + 1       # 1-based
COL_GTP_NO    = HEADERS.index("GTP No.") + 1
COL_ITEM_NO   = HEADERS.index("Item No.") + 1
COL_PRICE_A   = HEADERS.index("Price A (Rs/km)") + 1
COL_PRICE_B   = HEADERS.index("Price B (Rs/km)") + 1
COL_PRICE_C   = HEADERS.index("Price C (Rs/km)") + 1

BLUE        = "003366"
YELLOW_FILL = PatternFill("solid", fgColor="FFF59D")
ALT_FILL    = PatternFill("solid", fgColor="F5F5F5")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin", color="BBBBBB"),
    right=Side(style="thin", color="BBBBBB"),
    top=Side(style="thin", color="BBBBBB"),
    bottom=Side(style="thin", color="BBBBBB"),
)


def _hdr_style(cell):
    cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    cell.fill      = PatternFill("solid", fgColor=BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = THIN_BORDER


def _data_style(cell, alt: bool, is_margin: bool):
    cell.font      = Font(name="Calibri", size=10)
    cell.fill      = YELLOW_FILL if is_margin else (ALT_FILL if alt else WHITE_FILL)
    cell.alignment = Alignment(vertical="center")
    cell.border    = THIN_BORDER


def _load_wb() -> openpyxl.Workbook:
    if os.path.exists(REGISTRY_PATH):
        return openpyxl.load_workbook(REGISTRY_PATH)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GTP_Registry"
    ws.append(HEADERS)
    for col in range(1, len(HEADERS) + 1):
        _hdr_style(ws.cell(1, col))
    ws.freeze_panes = "A2"
    _set_col_widths(ws)
    wb.save(REGISTRY_PATH)
    return wb


def _set_col_widths(ws):
    widths = [18, 9, 28, 22, 12, 7, 11, 12, 11, 18, 18, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _find_row(ws, gtp_no: str, item_no: str) -> Optional[int]:
    for row in ws.iter_rows(min_row=2):
        if str(row[COL_GTP_NO - 1].value) == str(gtp_no) and \
           str(row[COL_ITEM_NO - 1].value) == str(item_no):
            return row[0].row
    return None


def get_margin(gtp_no: str, item_no: str) -> float:
    """Return the margin % for this item from the registry, or DEFAULT_MARGIN."""
    if not os.path.exists(REGISTRY_PATH):
        return DEFAULT_MARGIN
    wb = _load_wb()
    ws = wb["GTP_Registry"]
    row_num = _find_row(ws, gtp_no, item_no)
    if row_num is None:
        return DEFAULT_MARGIN
    val = ws.cell(row_num, COL_MARGIN).value
    try:
        return float(val)
    except (TypeError, ValueError):
        return DEFAULT_MARGIN


def upsert_row(gtp_no: str, item_no: str, cable: dict,
               prices: dict, margin_pct: float):
    """Insert or update a registry row. Preserves user-edited Margin % on update."""
    os.makedirs(os.path.dirname(REGISTRY_PATH), exist_ok=True)
    wb = _load_wb()
    ws = wb["GTP_Registry"]

    row_num = _find_row(ws, gtp_no, item_no)
    is_new  = row_num is None

    if is_new:
        row_num = ws.max_row + 1
        # New row — write margin_pct as default
        effective_margin = margin_pct
    else:
        # Existing row — keep whatever the user has set
        existing = ws.cell(row_num, COL_MARGIN).value
        try:
            effective_margin = float(existing)
        except (TypeError, ValueError):
            effective_margin = margin_pct

    alt = (row_num % 2 == 0)
    values = [
        gtp_no,
        item_no,
        cable.get("designation", ""),
        cable.get("config", ""),
        cable.get("conductor_material", ""),
        cable.get("num_cores", ""),
        cable.get("conductor_area_mm2", ""),
        cable.get("voltage_kv", ""),
        effective_margin,
        round(prices.get("A", 0), 0),
        round(prices.get("B", 0), 0),
        round(prices.get("C", 0), 0),
        datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"),
    ]

    for col, val in enumerate(values, 1):
        c = ws.cell(row_num, col, val)
        _data_style(c, alt, is_margin=(col == COL_MARGIN))
        if col in (COL_PRICE_A, COL_PRICE_B, COL_PRICE_C):
            c.number_format = '#,##0'
        if col == COL_MARGIN:
            c.number_format = '0.0'

    _set_col_widths(ws)
    wb.save(REGISTRY_PATH)
    return effective_margin
