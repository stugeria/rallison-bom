"""
RM Prices reader — loads material prices from ravin_bom_master.xlsx → RM_Prices sheet.
Falls back to data/rm_prices.json for any material not listed or with price = 0.

Procurement workflow:
  1. Open ravin_bom_master.xlsx
  2. Go to RM_Prices sheet
  3. Fill / update rm_price_per_kg column
  4. Save — next agent run picks up new prices automatically
"""

import json
import os
from typing import Optional

BASE_DIR   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(BASE_DIR, "ravin_bom_master.xlsx")
JSON_PATH  = os.path.join(BASE_DIR, "data", "rm_prices.json")


def load_rm_prices() -> dict:
    """
    Returns {material_code: price_per_kg}.
    Excel takes priority over JSON; JSON fills any gap.
    """
    # Load JSON baseline
    prices: dict = {}
    if os.path.exists(JSON_PATH):
        with open(JSON_PATH) as f:
            prices = json.load(f)

    # Override with Excel values (only where price > 0)
    excel_prices = _read_excel_prices()
    for code, price in excel_prices.items():
        if price and price > 0:
            prices[code] = price

    return prices


def _read_excel_prices() -> dict:
    try:
        import openpyxl
        if not os.path.exists(EXCEL_PATH):
            return {}
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        if "RM_Prices" not in wb.sheetnames:
            return {}
        ws = wb["RM_Prices"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {}
        headers = [str(h).strip().lower() if h else "" for h in rows[0]]
        try:
            code_col  = headers.index("material_code")
            price_col = headers.index("rm_price_per_kg")
        except ValueError:
            return {}
        result = {}
        for row in rows[1:]:
            if not row or not row[code_col]:
                continue
            code = str(row[code_col]).strip()
            try:
                price = float(row[price_col] or 0)
            except (TypeError, ValueError):
                continue
            if code and price > 0:
                result[code] = price
        return result
    except Exception:
        return {}
