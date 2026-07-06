"""
Drum / packing cost lookup.

Priority:
  1. Exact item entry in ravin_bom_master.xlsx → Drum_Costs sheet
     (matched by GTP_No + Item_No, or Cable_Family + Area range)
  2. % fallback from data/drum_rates.json applied to material_cost_per_km
"""

import json
import os
from typing import Optional

BASE_DIR  = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR  = os.path.join(BASE_DIR, "data")
EXCEL_PATH = os.path.join(BASE_DIR, "ravin_bom_master.xlsx")

_drum_rates_cache: dict = {}


def _load_drum_rates() -> list:
    global _drum_rates_cache
    if not _drum_rates_cache:
        path = os.path.join(DATA_DIR, "drum_rates.json")
        with open(path) as f:
            _drum_rates_cache = json.load(f)
    return _drum_rates_cache.get("rules", [])


def _read_excel_drum_costs() -> list[dict]:
    """Read Drum_Costs sheet from ravin_bom_master.xlsx. Returns list of row dicts."""
    try:
        import openpyxl
        if not os.path.exists(EXCEL_PATH):
            return []
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        if "Drum_Costs" not in wb.sheetnames:
            return []
        ws = wb["Drum_Costs"]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h).strip() if h else "" for h in rows[0]]
        result = []
        for row in rows[1:]:
            d = dict(zip(headers, row))
            # Skip instruction/blank rows — must have Cost_per_km filled and > 0
            try:
                cost = float(d.get("Cost_per_km") or 0)
            except (TypeError, ValueError):
                continue
            if cost <= 0:
                continue
            result.append(d)
        return result
    except Exception:
        return []


def lookup_drum_cost_per_km(
    material_cost_per_km: float,
    product_family: str,
    conductor_material: str,
    area_mm2: float,
    num_cores: float,
    drum_type: str,
    gtp_no: str = "",
    item_no: str = "",
) -> tuple[float, str]:
    """
    Returns (drum_cost_per_km, source) where source is "excel", "pct_rule", or "zero".
    drum_type: "wooden" | "steel" | "coil"
    """
    drum_type = (drum_type or "wooden").lower().strip()
    fam = (product_family or "").upper()
    mat = (conductor_material or "copper").lower()

    # ── 1. Excel exact match ──────────────────────────────────────────────────
    excel_rows = _read_excel_drum_costs()
    for row in excel_rows:
        gtp_match  = not row.get("GTP_No")  or str(row["GTP_No"]).strip() == str(gtp_no).strip()
        item_match = not row.get("Item_No") or str(row["Item_No"]).strip() == str(item_no).strip()
        fam_match  = not row.get("Cable_Family") or str(row["Cable_Family"]).upper() == fam
        area_match = True
        if row.get("Area_mm2"):
            try:
                area_match = abs(float(row["Area_mm2"]) - area_mm2) < 0.1
            except (TypeError, ValueError):
                pass
        if gtp_match and item_match and fam_match and area_match:
            cost = float(row.get("Cost_per_km", 0) or 0)
            return cost, "excel"

    # ── 2. % fallback from drum_rates.json ───────────────────────────────────
    pct = _lookup_drum_pct(fam, mat, area_mm2, num_cores, drum_type)
    if pct is not None:
        return round(material_cost_per_km * pct / 100, 2), f"pct_{pct}%"

    return 0.0, "zero"


def _lookup_drum_pct(fam: str, mat: str, area_mm2: float,
                     num_cores: float, drum_type: str) -> Optional[float]:
    for rule in _load_drum_rates():
        # Family filter
        families = [f.upper() for f in rule.get("families", [])]
        if families and fam not in families:
            continue
        # Conductor filter
        if "conductor" in rule and rule["conductor"].lower() != mat:
            continue
        # Core count filters
        if "max_cores" in rule and num_cores > rule["max_cores"]:
            continue
        if "min_cores" in rule and num_cores < rule["min_cores"]:
            continue
        # Drum type filter
        allowed_types = [t.lower() for t in rule.get("drum_types", [])]
        if allowed_types and drum_type not in allowed_types:
            continue

        # Flat rate rule
        if "drum_pct" in rule:
            return float(rule["drum_pct"])

        # Banded rule
        for band in rule.get("bands", []):
            if area_mm2 <= band.get("up_to_mm2", float("inf")):
                return float(band["drum_pct"])

    return None


def is_steel_drum(drum_type: str) -> bool:
    return (drum_type or "").lower().strip() == "steel"
