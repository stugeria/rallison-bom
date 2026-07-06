"""
Local BOM Detail Store — persists production and costing BOM weights to
output/bom_detail.xlsx with two sheets: BOM_Production and BOM_Costing.

Schema (one row per cable × layer × BOM type):
  BOM No. | GTP No. | BOM Type | Item No. | Item Name | Item Code |
  RM Code | RM Description | Weight (kg/km)

BOM No. format: {GTP_No}-{Item_No}-{A/B/C}
Dedup key: (BOM No., RM Description) — skips rows already written.
"""

import os
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError("openpyxl not installed — run: pip install openpyxl")

BOM_DETAIL_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "output", "bom_detail.xlsx"
)

SHEET_PRODUCTION = "BOM_Production"
SHEET_COSTING    = "BOM_Costing"

HEADERS = [
    "BOM No.", "GTP No.", "BOM Type", "Item No.", "Item Name", "Item Code",
    "RM Code", "RM Description", "Weight (kg/km)",
]

COL_WIDTHS = [28, 20, 10, 9, 28, 14, 14, 28, 16]

# RM Code prefix map by material_key prefix
_RM_PREFIX = {
    "copper_conductor":    ("RM-CU-001", "Copper Conductor"),
    "aluminium_conductor": ("RM-AL-001", "Aluminium Conductor"),
    "xlpe_insulation":     ("RM-IN-001", "XLPE Insulation Compound"),
    "pvc_insulation":      ("RM-IN-002", "PVC Insulation Compound"),
    "pvc_inner_sheath":    ("RM-SH-001", "PVC Inner Sheath Compound"),
    "pvc_outer_sheath":    ("RM-SH-002", "PVC Outer Sheath Compound"),
    "pvc_armoured_sheath": ("RM-SH-003", "PVC Armoured Sheath Compound"),
    "lszh_sheath":         ("RM-SH-004", "LSZH Sheath Compound"),
    "lszh_outer_sheath":   ("RM-SH-005", "LSZH Outer Sheath Compound"),
    "lszh_inner_sheath":   ("RM-SH-006", "LSZH Inner Sheath Compound"),
    "frlsh_sheath":        ("RM-SH-007", "FRLSH Sheath Compound"),
    "hffr_sheath":         ("RM-SH-008", "HFFR Sheath Compound"),
    "gs_flat_strip_armour":("RM-AR-001", "GI Flat Strip Armour"),
    "gs_round_wire_armour":("RM-AR-002", "GI Round Wire Armour"),
    "copper_tape_screen":  ("RM-SC-001", "Copper Tape Screen"),
    "copper_wire_screen":  ("RM-SC-002", "Copper Wire Screen"),
    "glass_mica_tape":     ("RM-TP-001", "Glass Mica Fire Barrier Tape"),
    "binder_tape":         ("RM-TP-002", "Binder Tape"),
    "binding_tape_pp":     ("RM-TP-003", "PP Binding Tape"),
    "petp_tape":           ("RM-TP-004", "PETP Tape"),
    "swelling_tape":       ("RM-TP-005", "Swelling Tape"),
    "pe_tape":             ("RM-TP-006", "PE Tape"),
    "al_mylar_pe_tape":    ("RM-TP-007", "Al-Mylar PE Tape"),
    "pp_filler":           ("RM-FL-001", "PP Filler"),
    "pvc_filler":          ("RM-FL-002", "PVC Filler"),
    "filler_compound":     ("RM-FL-003", "Filler Compound"),
    "rubber_epdm":         ("RM-IN-003", "EPDM Rubber Compound"),
    "drain_wire":          ("RM-DW-001", "Drain Wire"),
}

BLUE        = "003366"
ALT_FILL    = PatternFill("solid", fgColor="F5F5F5")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin",   color="BBBBBB"),
    right=Side(style="thin",  color="BBBBBB"),
    top=Side(style="thin",    color="BBBBBB"),
    bottom=Side(style="thin", color="BBBBBB"),
)


def _rm_info(material_key: str) -> tuple:
    return _RM_PREFIX.get(material_key, ("", material_key.replace("_", " ").title()))


def _hdr_style(cell):
    cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    cell.fill      = PatternFill("solid", fgColor=BLUE)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = THIN_BORDER


def _data_style(cell, alt: bool, is_weight: bool = False):
    cell.font      = Font(name="Calibri", size=10)
    cell.fill      = ALT_FILL if alt else WHITE_FILL
    cell.alignment = Alignment(vertical="center",
                               horizontal="right" if is_weight else "left")
    cell.border    = THIN_BORDER
    if is_weight:
        cell.number_format = '#,##0.00'


def _load_or_create_wb() -> openpyxl.Workbook:
    if os.path.exists(BOM_DETAIL_PATH):
        return openpyxl.load_workbook(BOM_DETAIL_PATH)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name in (SHEET_PRODUCTION, SHEET_COSTING):
        ws = wb.create_sheet(sheet_name)
        ws.append(HEADERS)
        for col in range(1, len(HEADERS) + 1):
            _hdr_style(ws.cell(1, col))
        for i, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A2"
    os.makedirs(os.path.dirname(BOM_DETAIL_PATH), exist_ok=True)
    wb.save(BOM_DETAIL_PATH)
    return wb


def _existing_bom_keys(ws) -> set:
    keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            keys.add((str(row[0]), str(row[7])))   # (BOM No., RM Description)
    return keys


def write_bom_rows(
    gtp_no: str,
    item_no: str,
    item_name: str,
    item_code: str,
    boms: dict,   # {"A": {"costing": [...], "production": [...]}, "B": ..., "C": ...}
):
    """
    Write all BOM type × mode rows to bom_detail.xlsx.
    Skips rows already present (dedup on BOM No. + RM Description).
    """
    os.makedirs(os.path.dirname(BOM_DETAIL_PATH), exist_ok=True)
    wb = _load_or_create_wb()

    ws_prod = wb[SHEET_PRODUCTION]
    ws_cost = wb[SHEET_COSTING]

    existing_prod = _existing_bom_keys(ws_prod)
    existing_cost = _existing_bom_keys(ws_cost)

    for bom_type in ("A", "B", "C"):
        bom_no = f"{gtp_no}-{item_no}-{bom_type}"
        for mode, ws, existing in (
            ("production", ws_prod, existing_prod),
            ("costing",    ws_cost, existing_cost),
        ):
            rows = boms.get(bom_type, {}).get(mode, [])
            for r in rows:
                mat_key = r.get("material", "")
                rm_code, rm_desc = _rm_info(mat_key)
                dedup_key = (bom_no, rm_desc)
                if dedup_key in existing:
                    continue
                existing.add(dedup_key)

                row_num = ws.max_row + 1
                alt     = (row_num % 2 == 0)
                values  = [
                    bom_no, gtp_no, bom_type, item_no,
                    item_name, item_code,
                    rm_code, rm_desc,
                    round(float(r.get("weight_kg_per_km", 0)), 3),
                ]
                for col, val in enumerate(values, 1):
                    c = ws.cell(row_num, col, val)
                    _data_style(c, alt, is_weight=(col == 9))

    wb.save(BOM_DETAIL_PATH)
