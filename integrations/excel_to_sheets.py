"""
Sync ravin_bom_master.xlsx → Google Sheets.

Reads all reference sheets from the Excel workbook and replicates them
to the configured Google Spreadsheet, preserving headers, data, and
basic formatting (bold headers, frozen rows).

Usage:
    python integrations/excel_to_sheets.py [--spreadsheet-id SHEET_ID]

If --spreadsheet-id is omitted, uses SPREADSHEET_ID from environment / settings.
"""

import argparse
import os
import sys
import time

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl>=3.1.0")
    sys.exit(1)

try:
    import gspread
    from google.oauth2.service_account import Credentials
except ImportError:
    print("ERROR: gspread not installed. Run: pip install gspread google-auth")
    sys.exit(1)

BASE_DIR   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(BASE_DIR, "ravin_bom_master.xlsx")

# Sheets to sync (in order). README is skipped — it contains prose, not data tables.
SYNC_SHEETS = [
    "Layer_Registry",
    "Cable_Families",
    "Master_Data",
    "Lay_Factors",
    "Extrusion_Tolerances",
    "RM_Prices",
    "Drum_Costs",
    "Margins",
]

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

# Google Sheets API allows max 60 writes/min on free tier — add a small delay
_WRITE_DELAY_S = 0.5


def _connect(spreadsheet_id: str):
    creds_file = os.environ.get(
        "GOOGLE_CREDENTIALS_FILE",
        os.path.expanduser("~/.config/ravin_cables/google_credentials.json")
    )
    if not os.path.exists(creds_file):
        raise FileNotFoundError(
            f"Google credentials not found at: {creds_file}\n"
            "Set GOOGLE_CREDENTIALS_FILE env var or place credentials at the path above."
        )
    creds = Credentials.from_service_account_file(creds_file, scopes=SCOPES)
    gc = gspread.authorize(creds)
    ss = gc.open_by_key(spreadsheet_id)
    return ss


def _get_or_create_worksheet(ss, title: str, rows: int = 200, cols: int = 20):
    try:
        ws = ss.worksheet(title)
    except gspread.exceptions.WorksheetNotFound:
        ws = ss.add_worksheet(title=title, rows=rows, cols=cols)
        print(f"    Created new worksheet: {title}")
    return ws


def _xl_sheet_to_rows(wb, sheet_name: str) -> list[list]:
    """Extract all non-empty rows from an Excel sheet as plain Python values."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        # Convert to strings/numbers, skipping fully-empty rows
        row_vals = [("" if v is None else v) for v in row]
        if any(v != "" for v in row_vals):
            rows.append(row_vals)
    return rows


def _sync_sheet(ss, title: str, data_rows: list[list]):
    """Clear the target worksheet and write new data."""
    if not data_rows:
        print(f"  ⚠ No data for sheet '{title}' — skipped")
        return

    ws = _get_or_create_worksheet(ss, title, rows=max(len(data_rows) + 10, 50), cols=max(len(data_rows[0]) + 2, 10))

    # Clear existing content
    ws.clear()
    time.sleep(_WRITE_DELAY_S)

    # Write all rows in one batch (gspread update supports 2D arrays)
    # Chunk into 500-row batches to stay under API size limits
    CHUNK = 500
    for start in range(0, len(data_rows), CHUNK):
        chunk = data_rows[start:start + CHUNK]
        end_row = start + len(chunk)
        end_col = max(len(r) for r in chunk)
        range_name = f"A{start + 1}:{_col_letter(end_col)}{end_row}"
        ws.update(range_name, chunk, value_input_option="USER_ENTERED")
        time.sleep(_WRITE_DELAY_S)

    # Bold + freeze the header row
    try:
        ws.format("1:1", {
            "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
            "backgroundColor": {"red": 0.0, "green": 0.2, "blue": 0.4},
        })
        ws.freeze(rows=1)
        time.sleep(_WRITE_DELAY_S)
    except Exception as e:
        print(f"    Warning: could not format header row — {e}")

    print(f"  ✓ {title}: {len(data_rows)} rows synced")


def _col_letter(n: int) -> str:
    """Convert 1-based column number to spreadsheet letter (e.g. 1→A, 27→AA)."""
    result = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result


def sync_all(spreadsheet_id: str):
    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: Excel workbook not found at {EXCEL_PATH}")
        print("Run: python tools/build_excel.py")
        sys.exit(1)

    print(f"Loading Excel workbook: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)  # data_only=True reads cell values, not formulas

    print(f"Connecting to Google Sheets (ID: {spreadsheet_id})")
    ss = _connect(spreadsheet_id)
    print(f"Connected to: {ss.title}\n")

    for sheet_name in SYNC_SHEETS:
        print(f"Syncing: {sheet_name}")
        rows = _xl_sheet_to_rows(wb, sheet_name)
        _sync_sheet(ss, sheet_name, rows)

    print(f"\nDone. {len(SYNC_SHEETS)} sheets synced to Google Sheets.")
    print(f"Spreadsheet URL: https://docs.google.com/spreadsheets/d/{spreadsheet_id}")


def main():
    parser = argparse.ArgumentParser(description="Sync ravin_bom_master.xlsx → Google Sheets")
    parser.add_argument("--spreadsheet-id", help="Google Sheets spreadsheet ID (overrides env var)")
    args = parser.parse_args()

    spreadsheet_id = args.spreadsheet_id or os.environ.get("SPREADSHEET_ID")
    if not spreadsheet_id:
        # Try loading from settings
        sys.path.insert(0, BASE_DIR)
        try:
            from config.settings import SPREADSHEET_ID
            spreadsheet_id = SPREADSHEET_ID
        except ImportError:
            pass

    if not spreadsheet_id:
        print("ERROR: No spreadsheet ID provided.")
        print("Usage: python integrations/excel_to_sheets.py --spreadsheet-id <ID>")
        print("   or: set SPREADSHEET_ID environment variable")
        sys.exit(1)

    sync_all(spreadsheet_id)


if __name__ == "__main__":
    main()
