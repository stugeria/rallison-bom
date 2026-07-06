"""
BOM Walkthrough — step-by-step calculation trace for 3 cable items.
Prints every formula with inputs/outputs so you can verify each step,
then writes an Excel BOM file.
"""

import sys, os, math, json
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from core.gtp_parser import parse_gtp
from core.bom_calculator import (
    RESISTIVITY,
    get_density,
    get_conductor_lay_factor,
    get_cabling_lay_factor,
    get_thickness_tolerance_factor,
    _load_data,
)

GTP_PDF = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "2-23077-GTP.pdf")
OUT_EXCEL = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                         "output", "BOM_3Items_Walkthrough.xlsx")

DIVIDER = "=" * 80
SUBDIV  = "-" * 60

_load_data()  # pre-load JSON data files


# ── helpers ──────────────────────────────────────────────────────────────────

def _tol_band(t_mm: float) -> str:
    if t_mm < 1.0:   return "thin  (< 1 mm)"
    if t_mm < 2.0:   return "medium (1–2 mm)"
    return             "thick  (≥ 2 mm)"


def step(n: int, title: str):
    print(f"\n  Step {n}: {title}")
    print(f"  {'·'*56}")


def result(label: str, value, unit: str = ""):
    print(f"    → {label}: {value} {unit}".rstrip())


# ── per-layer calculators (verbose) ──────────────────────────────────────────

def calc_conductor_verbose(cable: dict, bom_type: str, phase: str = "") -> list:
    """Returns list of BOM row dicts; prints step-by-step trace."""
    tag = f" ({phase})" if phase else ""

    # Pick the right R_dc
    if phase == "Neutral":
        r_dc = cable["neutral_dc_resistance_ohm_per_km"]
        label = "Conductor (Neutral)"
    else:
        r_dc  = cable["dc_resistance_ohm_per_km"]
        label = "Conductor (Phase)" if phase == "Phase" else "Conductor"

    mat  = cable["conductor_material"]          # "copper" or "aluminium"
    nw   = cable.get("num_wires", 7)
    n_ph = 3 if phase == "Phase" else (1 if phase == "Neutral" else math.ceil(cable.get("num_cores", 1)))

    rho       = RESISTIVITY[mat]
    r_dc_m    = r_dc / 1000.0
    area_mm2  = rho / r_dc_m

    density    = get_density(f"{mat}_conductor", bom_type)
    lay_factor = get_conductor_lay_factor(nw, bom_type)

    wt_single = area_mm2 * density * lay_factor          # kg/km per core
    wt_total  = round(wt_single * n_ph, 3)

    # Print trace
    step(1, f"Conductor{tag} ({mat.title()}, Class 2, {bom_type})")
    print(f"    GTP input   : R_dc = {r_dc} Ω/km, num_wires = {nw}, num_cores = {n_ph}")
    print(f"    Resistivity : ρ = 1/58 = {rho:.6f} Ω·mm²/m  (copper = 1/58, aluminium = 1/35)")
    print(f"    R_dc per m  : {r_dc} / 1000 = {r_dc_m:.6f} Ω/m")
    print(f"    Cross-section: area = ρ / R_dc_m = {rho:.6f} / {r_dc_m:.6f}")
    result("Conductor area", f"{area_mm2:.4f}", "mm²")
    print(f"    Density     : {density} g/cm³  →  (= kg/dm³, same numeric value in mm²·km units)")
    print(f"    Lay factor  : {lay_factor}  (num_wires={nw} → band 1–7)")
    print(f"    Weight/core : {area_mm2:.4f} × {density} × {lay_factor} = {wt_single:.4f} kg/km")
    print(f"    × {n_ph} core(s) = {wt_total} kg/km")

    mat_key = f"{mat}_conductor"
    od_approx = round(1.13 * math.sqrt(area_mm2), 2)
    return [{
        "layer": label, "material": mat_key,
        "effective_area_mm2": round(area_mm2, 4), "lay_factor": lay_factor,
        "density_g_cm3": density, "num_cores": n_ph,
        "weight_kg_per_km": wt_total,
        "_od_mm": od_approx,   # internal: used to seed OD chain
    }]


def calc_annular_verbose(layer: dict, id_mm: float, bom_type: str, num_cores: int,
                         step_n: int, apply_cabling_lay: bool = False,
                         tag: str = "") -> dict:
    mat_key    = layer["material_key"]
    layer_name = layer["layer_name"] + (f" ({tag})" if tag else "")
    t_nom      = layer["nominal_thickness_mm"]
    t_type     = layer.get("thickness_type") or "Nominal"
    density    = get_density(mat_key, bom_type)
    tol_factor = get_thickness_tolerance_factor(t_nom, t_type, bom_type)
    t_eff      = t_nom * tol_factor
    od_mm      = round(id_mm + 2 * t_eff, 3)
    lay_factor = get_cabling_lay_factor(bom_type) if apply_cabling_lay else 1.0
    ann_area   = (math.pi / 4) * (od_mm**2 - id_mm**2)   # mm²
    wt         = round(ann_area * density * lay_factor * num_cores, 3)

    step(step_n, f"{layer_name} — annular extrusion ({bom_type})")
    print(f"    GTP input   : nominal thickness = {t_nom} mm, type = {t_type}, num_cores = {num_cores}")
    print(f"    Band        : {_tol_band(t_nom)}")
    print(f"    Tol factor  : {tol_factor}  (table: Nominal/thin→1.10, medium→1.08, thick→1.05;")
    print(f"                              Minimum/thin→1.15, medium→1.12, thick→1.08, costing mode)")
    print(f"    Eff. thick  : {t_nom} × {tol_factor} = {t_eff:.3f} mm")
    print(f"    Inner OD    : {id_mm:.3f} mm  →  Outer OD = {id_mm:.3f} + 2×{t_eff:.3f} = {od_mm} mm")
    print(f"    Ann. area   : π/4 × ({od_mm}² − {id_mm:.3f}²) = {ann_area:.4f} mm²")
    print(f"    Density     : {density} g/cm³")
    print(f"    Lay factor  : {lay_factor}  (cabling lay applied: {apply_cabling_lay})")
    print(f"    Weight/km   : {ann_area:.4f} × {density} × {lay_factor} × {num_cores} core(s)")
    result("Weight", wt, "kg/km")

    return {
        "layer": layer_name, "material": mat_key,
        "id_mm": round(id_mm, 3), "effective_thickness_mm": round(t_eff, 3),
        "od_mm": od_mm, "tolerance_factor": tol_factor,
        "density_g_cm3": density, "lay_factor": lay_factor,
        "weight_kg_per_km": wt,
    }


# ── main walkthrough ──────────────────────────────────────────────────────────

def run_walkthrough(bom_type: str = "costing") -> list:
    print(f"\n{DIVIDER}")
    print(f"  BOM WALKTHROUGH  —  GTP: SM-121124-0  |  Mode: {bom_type.upper()}")
    print(DIVIDER)

    gtp = parse_gtp(GTP_PDF)
    cables = gtp["cables"][:3]

    all_results = []

    for cable in cables:
        item   = cable.get("item_no")
        cfg    = cable.get("config", "")
        desig  = cable.get("designation", "")
        n_cores = cable.get("num_cores", 1)
        neutral_rdc = cable.get("neutral_dc_resistance_ohm_per_km")
        is_35c = (n_cores == 3.5) and neutral_rdc is not None

        print(f"\n{DIVIDER}")
        print(f"  ITEM {item}:  {cfg}  {desig}  (GTP Type A, {bom_type.upper()} BOM)")
        print(DIVIDER)
        print(f"  From GTP:")
        print(f"    conductor_material   : {cable['conductor_material']}")
        print(f"    dc_resistance        : {cable['dc_resistance_ohm_per_km']} Ω/km")
        if is_35c:
            print(f"    neutral_dc_resistance: {neutral_rdc} Ω/km  (3.5C split)")
        print(f"    num_wires            : {cable.get('num_wires', 7)}")
        print(f"    num_cores            : {n_cores}")
        print(f"    Layers from GTP:     :", [l['layer_name'] for l in cable.get('layers', [])])

        bom_rows = []
        step_n = 0

        # ── Conductor(s) ──────────────────────────────────────────────────────
        if is_35c:
            # Phase cores
            step_n += 1
            rows_ph = calc_conductor_verbose(cable, bom_type, phase="Phase")
            bom_rows.extend(rows_ph)
            # Neutral core
            neutral_cable = dict(cable)
            neutral_cable["dc_resistance_ohm_per_km"] = neutral_rdc
            step_n += 1
            rows_n = calc_conductor_verbose(neutral_cable, bom_type, phase="Neutral")
            bom_rows.extend(rows_n)
            # OD for chain: use phase core OD (phase is larger)
            current_od = rows_ph[0]["_od_mm"]
            effective_cores_insulation = 3   # 3 phase + 1 neutral handled per-layer
        else:
            step_n += 1
            rows = calc_conductor_verbose(cable, bom_type)
            bom_rows.extend(rows)
            current_od = rows[0]["_od_mm"]
            effective_cores_insulation = math.ceil(n_cores)

        print(f"\n    (Conductor OD estimated: 1.13 × √area = {current_od} mm — used as ID for next layer)")

        # ── Remaining layers ──────────────────────────────────────────────────
        _CABLING_LAY_KEYS = {
            "frlsh_outer_sheath", "pvc_outer_sheath", "pvc_frlsh_sheath",
            "bedding", "pvc_armoured_sheath", "pvc_inner_sheath",
        }
        _PER_CORE_KEYS = {
            "xlpe_insulation", "pvc_insulation",
            "conductor_screen", "insulation_screen",
        }

        for layer in cable.get("layers", []):
            mat_key = layer["material_key"]
            apply_lay = mat_key in _CABLING_LAY_KEYS
            step_n += 1

            if is_35c and mat_key in _PER_CORE_KEYS:
                # Phase insulation
                neutral_t = layer.get("neutral_nominal_thickness_mm")
                row_ph = calc_annular_verbose(
                    layer, current_od, bom_type, num_cores=3,
                    step_n=step_n, apply_cabling_lay=apply_lay, tag="Phase"
                )
                bom_rows.append(row_ph)
                step_n += 1
                # Neutral insulation — build a modified layer dict
                neutral_layer = dict(layer)
                neutral_layer["nominal_thickness_mm"] = neutral_t
                row_n = calc_annular_verbose(
                    neutral_layer, current_od, bom_type, num_cores=1,
                    step_n=step_n, apply_cabling_lay=apply_lay, tag="Neutral"
                )
                bom_rows.append(row_n)
                current_od = row_ph["od_mm"]   # advance by phase (larger)
            else:
                nc = effective_cores_insulation if mat_key in _PER_CORE_KEYS else 1
                row = calc_annular_verbose(
                    layer, current_od, bom_type, num_cores=nc,
                    step_n=step_n, apply_cabling_lay=apply_lay
                )
                bom_rows.append(row)
                current_od = row["od_mm"]

        # Summary
        total_wt = round(sum(r["weight_kg_per_km"] for r in bom_rows), 3)
        print(f"\n  {SUBDIV}")
        print(f"  SUMMARY — Item {item}: {cfg} {desig}")
        print(f"  {'Layer':<35} {'Material':<25} {'kg/km':>10}")
        print(f"  {SUBDIV}")
        for r in bom_rows:
            print(f"  {r['layer']:<35} {r['material']:<25} {r['weight_kg_per_km']:>10.3f}")
        print(f"  {SUBDIV}")
        print(f"  {'TOTAL':<35} {'':<25} {total_wt:>10.3f} kg/km")

        all_results.append({
            "item_no": item, "config": cfg, "designation": desig,
            "bom_type": bom_type, "bom_rows": bom_rows, "total_kg_per_km": total_wt,
        })

    return all_results


def write_excel(results_costing: list, results_production: list):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("openpyxl not installed — skipping Excel output")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM_3Items"

    # Styles
    hdr_font   = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill   = PatternFill("solid", fgColor="1F3864")
    item_fill  = PatternFill("solid", fgColor="D6E4F0")
    tot_fill   = PatternFill("solid", fgColor="BDD7EE")
    thin_side  = Side(style="thin")
    thin_bdr   = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center     = Alignment(horizontal="center", vertical="center")
    left_al    = Alignment(horizontal="left",   vertical="center")

    def hdr(ws, row, col, val, width=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font   = hdr_font
        c.fill   = hdr_fill
        c.alignment = center
        c.border = thin_bdr
        return c

    def cell(ws, row, col, val, fill=None, bold=False, num_fmt=None, align=None):
        c = ws.cell(row=row, column=col, value=val)
        c.border = thin_bdr
        if fill:    c.fill = fill
        if bold:    c.font = Font(bold=True)
        if num_fmt: c.number_format = num_fmt
        c.alignment = align or left_al
        return c

    # Title
    ws.merge_cells("A1:L1")
    t = ws["A1"]
    t.value = "RAVIN CABLES — BOM (3 Items, GTP SM-121124-0-A)"
    t.font  = Font(bold=True, size=13, color="1F3864")
    t.alignment = center
    ws.row_dimensions[1].height = 22

    # Column headers (row 2)
    headers = [
        "Item No.", "Config", "Designation", "Layer", "Material",
        "Inner OD\n(mm)", "Eff. Thickness\n(mm)", "Outer OD\n(mm)",
        "Density\n(g/cm³)", "Lay Factor", "Weight Costing\n(kg/km)", "Weight Production\n(kg/km)"
    ]
    col_widths = [8, 16, 22, 28, 24, 12, 14, 12, 12, 10, 16, 18]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        hdr(ws, 2, ci, h)
        ws.column_dimensions[ws.cell(2, ci).column_letter].width = w
    ws.row_dimensions[2].height = 32

    # Build lookup for production weights
    prod_lookup: dict[int, dict[str, float]] = {}
    for res in results_production:
        iid = res["item_no"]
        prod_lookup[iid] = {r["layer"]: r["weight_kg_per_km"] for r in res["bom_rows"]}

    row_n = 3
    for res in results_costing:
        iid  = res["item_no"]
        rows = res["bom_rows"]
        n_rows = len(rows)
        prod_map = prod_lookup.get(iid, {})

        for ri, r in enumerate(rows):
            is_first = (ri == 0)
            if is_first:
                # merge item/config/designation columns for this item block
                ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n + n_rows - 1, end_column=1)
                ws.merge_cells(start_row=row_n, start_column=2, end_row=row_n + n_rows - 1, end_column=2)
                ws.merge_cells(start_row=row_n, start_column=3, end_row=row_n + n_rows - 1, end_column=3)
                cell(ws, row_n, 1, iid,           fill=item_fill, bold=True, align=center)
                cell(ws, row_n, 2, res["config"],  fill=item_fill, bold=True, align=center)
                cell(ws, row_n, 3, res["designation"], fill=item_fill, bold=True)

            cell(ws, row_n, 4, r["layer"])
            cell(ws, row_n, 5, r["material"])
            cell(ws, row_n, 6, r.get("id_mm", "—"), num_fmt="0.000")
            cell(ws, row_n, 7, r.get("effective_thickness_mm", "—"), num_fmt="0.000")
            cell(ws, row_n, 8, r.get("od_mm", "—"), num_fmt="0.000")
            cell(ws, row_n, 9, r["density_g_cm3"], num_fmt="0.000")
            cell(ws, row_n, 10, r.get("lay_factor", 1.0), num_fmt="0.000")
            cell(ws, row_n, 11, r["weight_kg_per_km"], num_fmt="0.000")
            prod_wt = prod_map.get(r["layer"], "—")
            cell(ws, row_n, 12, prod_wt, num_fmt="0.000" if isinstance(prod_wt, float) else "@")
            row_n += 1

        # TOTAL row
        total_c = res["total_kg_per_km"]
        total_p = round(sum(prod_lookup.get(iid, {}).values()), 3)
        ws.merge_cells(start_row=row_n, start_column=1, end_row=row_n, end_column=10)
        tc = ws.cell(row=row_n, column=1, value=f"TOTAL — Item {iid}: {res['config']} {res['designation']}")
        tc.font = Font(bold=True); tc.fill = tot_fill; tc.alignment = center; tc.border = thin_bdr
        cell(ws, row_n, 11, total_c, fill=tot_fill, bold=True, num_fmt="0.000")
        cell(ws, row_n, 12, total_p, fill=tot_fill, bold=True, num_fmt="0.000")
        row_n += 2   # blank gap between items

    # Freeze header rows
    ws.freeze_panes = "A3"

    os.makedirs(os.path.dirname(OUT_EXCEL), exist_ok=True)
    wb.save(OUT_EXCEL)
    print(f"\n[Excel] BOM saved → {OUT_EXCEL}")


# ── entry point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    results_c = run_walkthrough("costing")
    results_p = run_walkthrough("production")
    write_excel(results_c, results_p)
    print("\nDone.")
