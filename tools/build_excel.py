"""
Build the Ravin Cables BOM Master Workbook (ravin_bom_master.xlsx).

Generates a standalone Excel workbook with 11 sheets:
  README, Layer_Registry, Cable_Families, Master_Data, Lay_Factors,
  Extrusion_Tolerances, RM_Prices, Drum_Costs, Margins,
  BOM_Calculator, Costing_Calculator

Run: python tools/build_excel.py
"""

import json
import os
import sys
from datetime import date

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl>=3.1.0")
    sys.exit(1)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")
OUTPUT_PATH = os.path.join(BASE_DIR, "ravin_bom_master.xlsx")

# ── Style constants ──────────────────────────────────────────────────────────
RAVIN_BLUE   = "003366"
RAVIN_LIGHT  = "D6E4F0"
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL  = PatternFill("solid", fgColor=RAVIN_BLUE)
SUBHDR_FONT  = Font(name="Calibri", bold=True, color=RAVIN_BLUE, size=10)
SUBHDR_FILL  = PatternFill("solid", fgColor=RAVIN_LIGHT)
BODY_FONT    = Font(name="Calibri", size=10)
ALT_FILL     = PatternFill("solid", fgColor="F2F7FC")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)
WRAP_ALIGN   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
THIN_BORDER  = Border(
    left=Side(style="thin",   color="AAAAAA"),
    right=Side(style="thin",  color="AAAAAA"),
    top=Side(style="thin",    color="AAAAAA"),
    bottom=Side(style="thin", color="AAAAAA"),
)


def _style_header_row(ws, row: int, max_col: int):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def _style_data_row(ws, row: int, max_col: int, alternate: bool = False):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        if alternate:
            cell.fill = ALT_FILL


def _set_col_widths(ws, widths: list):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _freeze(ws, cell: str):
    ws.freeze_panes = cell


# ── Sheet builders ───────────────────────────────────────────────────────────

def build_readme(wb):
    ws = wb.create_sheet("README")
    ws.sheet_properties.tabColor = "003366"
    ws["A1"] = "RAVIN CABLES LTD — BOM Master Workbook"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color=RAVIN_BLUE)
    ws["A2"] = f"Version: {date.today().isoformat()}   |   DO NOT share externally — contains pricing data"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="FF0000")

    ws["A4"] = "Sheet Index"
    ws["A4"].font = SUBHDR_FONT
    index = [
        ("Layer_Registry",      "All layer types, formula definitions, required inputs"),
        ("Cable_Families",      "9 cable family definitions with layer sequences"),
        ("Master_Data",         "Material densities (g/cm³), resistivity constants, design factors"),
        ("Lay_Factors",         "Conductor / cabling / armour lay factor lookup tables"),
        ("Extrusion_Tolerances","Thickness tolerance factors by band and type (Nominal/Minimum)"),
        ("RM_Prices",           "Current raw material prices (₹/kg) — update regularly"),
        ("Drum_Costs",          "Drum cost lookup by product type, conductor area, drum type"),
        ("Margins",             "Target margin % by product family and conductor area range"),
        ("BOM_Calculator",      "Interactive BOM calculator — enter cable specs, get weight per layer"),
        ("Costing_Calculator",  "Applies RM prices to BOM → floor price and selling price"),
    ]
    ws["A5"] = "Sheet"
    ws["B5"] = "Purpose"
    ws["A5"].font = SUBHDR_FONT
    ws["B5"].font = SUBHDR_FONT
    for i, (sheet, purpose) in enumerate(index, start=6):
        ws[f"A{i}"] = sheet
        ws[f"B{i}"] = purpose
        ws[f"A{i}"].font = Font(name="Calibri", bold=True, size=10)
        ws[f"B{i}"].font = BODY_FONT

    ws["A17"] = "Formula Reference"
    ws["A17"].font = SUBHDR_FONT
    formulas = [
        ("Conductor Weight",          "area = ρ/R_DC × resistance_factor   |   weight = area × density × lay_factor × num_cores"),
        ("Annular Layer (Extrusion)", "OD = ID + 2×(thickness×tol_factor)   |   weight = (π/4)(OD²-ID²) × density × lay_factor"),
        ("GS Flat Strip Armour",      "n = π(D+t)/(w+gap)   |   weight = n × w × t × density × lay × coverage"),
        ("GS Round Wire Armour",      "n = π(D+d)/(d+gap)   |   weight = n × (π/4×d²) × density × lay × coverage"),
        ("Copper Tape Screen",        "weight = π × mean_OD × (1+overlap/100) × thickness × density × num_cores"),
        ("Tape Wrap (general)",       "weight = π × mean_OD × (1+overlap/100) × thickness × density × n_layers"),
        ("PP Filler",                 "weight = ((π/4×OD_cab²) - n×(π/4×OD_core²)) × fill_factor × density"),
        ("Drain Wire",                "weight = (π/4 × d²) × density_Cu × n_pairs"),
    ]
    for i, (name, formula) in enumerate(formulas, start=18):
        ws[f"A{i}"] = name
        ws[f"B{i}"] = formula
        ws[f"A{i}"].font = Font(name="Calibri", bold=True, size=10)
        ws[f"B{i}"].font = Font(name="Courier New", size=9)

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 90


def build_layer_registry(wb):
    ws = wb.create_sheet("Layer_Registry")
    headers = [
        "layer_key", "display_name", "formula_type", "required_inputs",
        "density_key", "tolerance_applicable", "lay_category", "apply_cabling_lay",
        "cable_families", "notes"
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    _freeze(ws, "A2")

    with open(os.path.join(DATA_DIR, "layer_registry.json")) as f:
        registry = json.load(f)

    for i, (key, info) in enumerate(registry.items()):
        if key == "_meta":
            continue
        row = [
            key,
            info.get("display_name", ""),
            info.get("formula_type", ""),
            ", ".join(info.get("required_inputs", [])),
            info.get("density_key", ""),
            "Yes" if info.get("tolerance_applicable") else "No",
            info.get("lay_category", ""),
            "Yes" if info.get("apply_cabling_lay") else "No",
            ", ".join(info.get("cable_families", [])),
            info.get("notes", ""),
        ]
        ws.append(row)
        alt = (i % 2 == 1)
        _style_data_row(ws, i + 2, len(headers), alt)
        ws.cell(i + 2, 10).alignment = WRAP_ALIGN

    widths = [22, 30, 22, 42, 22, 10, 16, 14, 60, 70]
    _set_col_widths(ws, widths)
    ws.row_dimensions[1].height = 18


def build_cable_families(wb):
    ws = wb.create_sheet("Cable_Families")
    headers = ["code", "name", "standard", "voltage_kv", "layer_sequence", "operations"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    _freeze(ws, "A2")

    _volt_map = {
        "LT_PVC_FLEX": "1.1", "LT_XLPE_UNARM": "1.1", "LT_XLPE_ARMOURED": "1.1",
        "HT_11KV_SCREENED": "11", "CTRL_PVC_UNARM": "1.1", "CTRL_PVC_ARMOURED": "1.1",
        "MV_XLPE_SCREENED": "22/33", "INSTR_SCREENED": "0.25", "FLEX_CLASS5": "0.3",
    }

    with open(os.path.join(DATA_DIR, "cable_families.json")) as f:
        families = json.load(f)["families"]

    for i, fam in enumerate(families):
        row = [
            fam["code"],
            fam["name"],
            fam["standard"],
            _volt_map.get(fam["code"], ""),
            " → ".join(fam["layers"]),
            " | ".join(fam["operations"]),
        ]
        ws.append(row)
        alt = (i % 2 == 1)
        _style_data_row(ws, i + 2, len(headers), alt)
        for col in (5, 6):
            ws.cell(i + 2, col).alignment = WRAP_ALIGN

    widths = [20, 36, 16, 10, 80, 80]
    _set_col_widths(ws, widths)
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 48


def build_master_data(wb):
    ws = wb.create_sheet("Master_Data")

    # ── Densities section ───────────────────────────────────────────────────
    ws["A1"] = "Material Densities (g/cm³)"
    ws["A1"].font = SUBHDR_FONT
    ws["A1"].fill = SUBHDR_FILL

    density_headers = ["material_code", "description", "density_costing", "density_production", "unit"]
    for col, h in enumerate(density_headers, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER_ALIGN
        c.border = THIN_BORDER
    _freeze(ws, "A3")

    with open(os.path.join(DATA_DIR, "material_densities.json")) as f:
        densities = json.load(f)

    _descriptions = {
        "copper_conductor": "Copper conductor (annealed)",
        "aluminium_conductor": "Aluminium conductor (1350)",
        "xlpe_insulation": "XLPE insulation compound",
        "pvc_insulation": "PVC insulation (control/flex)",
        "semicon_screen": "Semiconducting XLPE (screen)",
        "pvc_flexible": "PVC flexible compound",
        "pvc_armoured_sheath": "PVC armoured sheath",
        "pvc_outer_sheath": "PVC outer sheath",
        "pvc_inner_sheath": "PVC inner sheath",
        "frlsh_sheath": "FR-LSH compound sheath",
        "gs_flat_strip_armour": "GS flat strip armour (7.85)",
        "gs_round_wire_armour": "GS round wire armour (SWA/AWA)",
        "copper_tape_screen": "Copper tape screen",
        "copper_wire_screen": "Copper wire concentric screen",
        "filler_compound": "Filler compound / bedding",
        "binder_tape": "Binder tape",
        "binding_tape_pp": "Binding tape (PP/polyester)",
        "petp_tape": "PETP/Al foil tape (individual screen)",
        "pp_filler": "PP rope filler",
        "rubber_epdm": "EPDM/EPR rubber insulation",
        "swelling_tape": "Swelling/water-blocking tape",
    }

    for i, (code, vals) in enumerate(densities.items()):
        row_no = i + 3
        row = [code, _descriptions.get(code, ""), vals["costing"], vals["production"], vals.get("unit", "g/cm3")]
        for col, v in enumerate(row, start=1):
            c = ws.cell(row=row_no, column=col, value=v)
            c.font = BODY_FONT
            c.border = THIN_BORDER
        if i % 2 == 1:
            for col in range(1, 6):
                ws.cell(row_no, col).fill = ALT_FILL

    # ── Resistivity constants ────────────────────────────────────────────────
    offset = len(densities) + 5
    ws.cell(offset, 1).value = "Resistivity Constants"
    ws.cell(offset, 1).font = SUBHDR_FONT
    ws.cell(offset, 1).fill = SUBHDR_FILL

    for col, h in enumerate(["material", "resistivity_ohm_mm2_per_m", "formula", "note"], start=1):
        c = ws.cell(offset + 1, col, h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER

    res_data = [
        ("Copper",    1/58, "1/58",  "IEC 60228, 20°C"),
        ("Aluminium", 1/35, "1/35",  "IEC 60228, 20°C"),
    ]
    for i, (mat, val, formula, note) in enumerate(res_data):
        r = offset + 2 + i
        for col, v in enumerate([mat, round(val, 6), formula, note], start=1):
            c = ws.cell(r, col, v)
            c.font = BODY_FONT
            c.border = THIN_BORDER

    widths = [24, 32, 18, 22, 12]
    _set_col_widths(ws, widths)


def build_lay_factors(wb):
    ws = wb.create_sheet("Lay_Factors")
    _freeze(ws, "A2")

    # Conductor section
    ws["A1"] = "Conductor Lay Factors (Standard Class 2)"
    ws["A1"].font = SUBHDR_FONT
    ws["A1"].fill = SUBHDR_FILL

    cond_headers = ["category", "min_wires", "max_wires", "costing_value", "production_value"]
    for col, h in enumerate(cond_headers, start=1):
        c = ws.cell(2, col, h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = THIN_BORDER; c.alignment = CENTER_ALIGN

    with open(os.path.join(DATA_DIR, "lay_factors.json")) as f:
        lf = json.load(f)

    row_no = 3
    for band in lf["conductor"]:
        for col, v in enumerate(["conductor", band["min_wires"], band["max_wires"], band["costing"], band["production"]], start=1):
            c = ws.cell(row_no, col, v); c.font = BODY_FONT; c.border = THIN_BORDER
        row_no += 1

    row_no += 1
    ws.cell(row_no, 1).value = "Fine Wire Conductor Lay Factors (Class 5/6 Flexible)"
    ws.cell(row_no, 1).font = SUBHDR_FONT
    ws.cell(row_no, 1).fill = SUBHDR_FILL
    row_no += 1
    for col, h in enumerate(cond_headers, start=1):
        c = ws.cell(row_no, col, h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = THIN_BORDER; c.alignment = CENTER_ALIGN
    row_no += 1
    for band in lf.get("fine_wire_conductor", []):
        for col, v in enumerate(["fine_wire_conductor", band["min_wires"], band["max_wires"], band["costing"], band["production"]], start=1):
            c = ws.cell(row_no, col, v); c.font = BODY_FONT; c.border = THIN_BORDER
        row_no += 1

    row_no += 1
    ws.cell(row_no, 1).value = "Cabling / Armour Lay Factors"
    ws.cell(row_no, 1).font = SUBHDR_FONT
    ws.cell(row_no, 1).fill = SUBHDR_FILL
    row_no += 1
    for col, h in enumerate(["category", "costing_value", "production_value"], start=1):
        c = ws.cell(row_no, col, h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = THIN_BORDER; c.alignment = CENTER_ALIGN
    row_no += 1
    for cat in ("cabling", "armour", "round_wire_armour"):
        vals = lf.get(cat, {})
        if vals:
            for col, v in enumerate([cat, vals.get("costing", ""), vals.get("production", "")], start=1):
                c = ws.cell(row_no, col, v); c.font = BODY_FONT; c.border = THIN_BORDER
            row_no += 1

    _set_col_widths(ws, [24, 12, 12, 16, 18])


def build_extrusion_tolerances(wb):
    ws = wb.create_sheet("Extrusion_Tolerances")
    headers = ["thickness_type", "band", "min_mm", "max_mm", "costing_factor", "production_factor", "notes"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    _freeze(ws, "A2")

    rows = [
        ("Nominal", "thin",   0.0,  1.0,  1.10, 1.00, "e.g. semicon screen, thin tape"),
        ("Nominal", "medium", 1.0,  2.0,  1.08, 1.00, "e.g. thin insulation"),
        ("Nominal", "thick",  2.0, 99.0,  1.05, 1.00, "e.g. thick insulation, sheaths"),
        ("Minimum", "thin",   0.0,  1.0,  1.15, 1.05, "conservative: thin layers with Minimum spec"),
        ("Minimum", "medium", 1.0,  2.0,  1.12, 1.05, "e.g. IS 7098 insulation Minimum thickness"),
        ("Minimum", "thick",  2.0, 99.0,  1.08, 1.05, "e.g. thick sheaths with Minimum spec"),
    ]
    for i, row in enumerate(rows):
        ws.append(list(row))
        _style_data_row(ws, i + 2, len(headers), i % 2 == 1)

    _set_col_widths(ws, [16, 10, 10, 10, 16, 18, 40])


def build_rm_prices(wb):
    ws = wb.create_sheet("RM_Prices")
    headers = ["material_code", "description", "rm_price_per_kg", "unit", "last_updated", "source"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    _freeze(ws, "A2")

    defaults = [
        # ── Conductors ────────────────────────────────────────────────────────
        ("copper_conductor",      "Copper conductor (ETP grade)",        "0.00",  "₹/kg", str(date.today()), "MCX / supplier"),
        ("aluminium_conductor",   "Aluminium conductor (EC grade)",      "0.00",  "₹/kg", str(date.today()), "LME / supplier"),
        # ── Insulations ───────────────────────────────────────────────────────
        ("xlpe_insulation",       "XLPE insulation compound",            "0.00",  "₹/kg", str(date.today()), "Dow/BASF"),
        ("pvc_insulation",        "PVC insulation compound",             "0.00",  "₹/kg", str(date.today()), "Supplier"),
        ("semicon_screen",        "Semiconducting XLPE compound",        "0.00",  "₹/kg", str(date.today()), "Supplier"),
        # ── Fire barrier ──────────────────────────────────────────────────────
        ("glass_mica_tape",       "Glass mica fire barrier tape",        "0.00",  "₹/kg", str(date.today()), "Supplier"),
        # ── Sheaths ───────────────────────────────────────────────────────────
        ("pvc_flexible",          "PVC flexible compound",               "0.00",  "₹/kg", str(date.today()), "Supplier"),
        ("pvc_armoured_sheath",   "PVC armoured/bedding compound",       "0.00",  "₹/kg", str(date.today()), "Supplier"),
        ("pvc_outer_sheath",      "PVC outer sheath compound",           "0.00",  "₹/kg", str(date.today()), "Supplier"),
        ("pvc_inner_sheath",      "PVC inner sheath compound",           "0.00",  "₹/kg", str(date.today()), "Supplier"),
        ("lszh_outer_sheath",     "LSZH outer sheath compound",          "0.00",  "₹/kg", str(date.today()), "Supplier"),
        ("lszh_inner_sheath",     "LSZH inner sheath compound",          "0.00",  "₹/kg", str(date.today()), "Supplier"),
        ("frlsh_sheath",          "FR-LSH compound",                     "0.00",  "₹/kg", str(date.today()), "Supplier"),
        ("hffr_sheath",           "HFFR compound",                       "0.00",  "₹/kg", str(date.today()), "Supplier"),
        ("lszh_sheath",           "LSZH general compound",               "0.00",  "₹/kg", str(date.today()), "Supplier"),
        # ── Armour ────────────────────────────────────────────────────────────
        ("gs_flat_strip_armour",  "GS flat strip armour (IS 3975)",      "0.00",  "₹/kg", str(date.today()), "Steel market"),
        ("gs_round_wire_armour",  "GS round wire armour (IS 3975)",      "0.00",  "₹/kg", str(date.today()), "Steel market"),
        # ── Screens ───────────────────────────────────────────────────────────
        ("copper_tape_screen",    "Copper tape screen (annealed)",       "0.00",  "₹/kg", str(date.today()), "MCX / supplier"),
        ("copper_wire_screen",    "Copper wire concentric screen",       "0.00",  "₹/kg", str(date.today()), "MCX / supplier"),
        # ── Fillers & tapes ───────────────────────────────────────────────────
        ("filler_compound",       "Filler/bedding compound",             "0.00",  "₹/kg", str(date.today()), "Supplier"),
        ("pp_filler",             "PP rope filler",                      "0.00",  "₹/kg", str(date.today()), "Supplier"),
        ("pvc_filler",            "PVC filler compound",                 "0.00",  "₹/kg", str(date.today()), "Supplier"),
        ("binder_tape",           "Binder tape",                         "0.00",  "₹/kg", str(date.today()), "Supplier"),
        ("binding_tape_pp",       "Binding tape (PP)",                   "0.00",  "₹/kg", str(date.today()), "Supplier"),
        ("petp_tape",             "PETP/Al foil tape",                   "0.00",  "₹/kg", str(date.today()), "Supplier"),
        ("swelling_tape",         "Swelling/water-blocking tape",        "0.00",  "₹/kg", str(date.today()), "Supplier"),
        ("rubber_epdm",           "EPDM/EPR rubber compound",            "0.00",  "₹/kg", str(date.today()), "Supplier"),
    ]
    for i, row in enumerate(defaults):
        ws.append(list(row))
        _style_data_row(ws, i + 2, len(headers), i % 2 == 1)

    note_cell = ws.cell(len(defaults) + 3, 1)
    note_cell.value = "⚠ Fill rm_price_per_kg column with current prices before running costing calculations."
    note_cell.font = Font(name="Calibri", bold=True, size=10, color="CC0000")

    _set_col_widths(ws, [24, 36, 16, 8, 14, 20])


def build_drum_costs(wb):
    ws = wb.create_sheet("Drum_Costs")
    headers = [
        "GTP_No", "Item_No", "Description",
        "Cable_Family", "Conductor", "Area_mm2",
        "Delivery_Length_m", "Drum_Type", "Cost_per_km"
    ]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    _freeze(ws, "A2")

    # Instruction rows
    instructions = [
        ["HOW TO USE THIS SHEET", "", "", "", "", "", "", "", ""],
        ["• Add one row per cable item where you know the exact drum cost.", "", "", "", "", "", "", "", ""],
        ["• GTP_No + Item_No uniquely identify the cable. Leave blank for a family/area default.", "", "", "", "", "", "", ""],
        ["• Cost_per_km = total drum/packing cost in ₹ per km of cable.", "", "", "", "", "", "", ""],
        ["• If an item is NOT listed here, the system uses % fallback rates from data/drum_rates.json.", "", "", "", "", "", "", ""],
        ["• Drum_Type: wooden / steel / coil  (steel triggers an alert in Telegram)", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", ""],
        ["--- FILL YOUR DATA BELOW THIS LINE ---", "", "", "", "", "", "", "", ""],
    ]
    RED  = Font(name="Calibri", size=9, color="CC0000", italic=True)
    GREY = Font(name="Calibri", size=9, color="666666", italic=True)
    for i, row_data in enumerate(instructions, start=2):
        for col, val in enumerate(row_data, start=1):
            c = ws.cell(i, col, val)
            c.font = RED if "HOW TO USE" in str(val) or "---" in str(val) else GREY
        ws.row_dimensions[i].height = 14

    # Example placeholder row (greyed out)
    ex_row = ws.max_row + 1
    example = ["IS 17505 (2) (2)", "1", "3.5C x 300 SQMM", "FIRE_SURVIVAL", "copper", 300, 500, "wooden", 0]
    for col, val in enumerate(example, start=1):
        c = ws.cell(ex_row, col, val)
        c.font = Font(name="Calibri", size=10, color="AAAAAA", italic=True)
        c.fill = PatternFill("solid", fgColor="F9F9F9")

    _set_col_widths(ws, [22, 10, 32, 20, 12, 12, 18, 12, 14])


def build_margins(wb):
    ws = wb.create_sheet("Margins")
    headers = ["product_family", "min_area_mm2", "max_area_mm2", "margin_pct", "notes"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    _freeze(ws, "A2")

    defaults = [
        ("LT_PVC_FLEX",      0,    999, 15, "LT PVC flexible"),
        ("LT_XLPE_UNARM",    0,    999, 15, "LT XLPE unarmoured"),
        ("LT_XLPE_ARMOURED", 0,    999, 15, "LT XLPE armoured"),
        ("HT_11KV_SCREENED", 0,    999, 18, "HT 11kV screened"),
        ("MV_XLPE_SCREENED", 0,    999, 20, "MV 22/33kV screened"),
        ("CTRL_PVC_UNARM",   0,    999, 15, "Control unarmoured"),
        ("CTRL_PVC_ARMOURED",0,    999, 15, "Control armoured"),
        ("INSTR_SCREENED",   0,    999, 18, "Instrumentation"),
        ("FLEX_CLASS5",      0,    999, 12, "Flexible Class 5/6"),
    ]
    for i, row in enumerate(defaults):
        ws.append(list(row))
        _style_data_row(ws, i + 2, len(headers), i % 2 == 1)

    _set_col_widths(ws, [22, 14, 14, 12, 30])


def build_bom_calculator(wb):
    """
    BOM Calculator sheet — user enters cable specs, Excel calculates BOM.
    Uses VLOOKUP against Master_Data, Lay_Factors, Extrusion_Tolerances.
    """
    ws = wb.create_sheet("BOM_Calculator")
    ws.sheet_properties.tabColor = "1F7A1F"

    # ── Input area ───────────────────────────────────────────────────────────
    ws["A1"] = "CABLE BOM CALCULATOR"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color=RAVIN_BLUE)

    inputs = [
        ("B3",  "Cable Designation:",          "E3",  "e.g. A2XFY-FRLSH"),
        ("B4",  "Cable Config:",                "E4",  "e.g. 3.5C x 70mm²"),
        ("B5",  "Conductor Material:",          "E5",  "copper / aluminium"),
        ("B6",  "DC Resistance (Ω/km):",        "E6",  "from GTP"),
        ("B7",  "Number of Wires:",             "E7",  "e.g. 19"),
        ("B8",  "Conductor OD (mm):",           "E8",  "from GTP or calculated"),
        ("B9",  "Number of Cores:",             "E9",  "e.g. 3.5"),
        ("B10", "GTP Type (A/B/C):",            "E10", "default A"),
        ("B11", "BOM Type (costing/production)","E11", "costing or production"),
        ("B12", "Cable Family Code:",           "E12", "e.g. LT_XLPE_ARMOURED"),
    ]
    for label_cell, label_val, input_cell, hint in inputs:
        ws[label_cell] = label_val
        ws[label_cell].font = Font(name="Calibri", bold=True, size=10)
        ws[input_cell] = hint
        ws[input_cell].font = Font(name="Calibri", italic=True, size=10, color="888888")

    ws["B14"] = "Layer-by-Layer BOM Results"
    ws["B14"].font = SUBHDR_FONT
    ws["B14"].fill = SUBHDR_FILL

    bom_headers = [
        "Layer No.", "Layer Name", "Material Key", "Formula Type",
        "Inner OD (mm)", "Thickness (mm)", "Outer OD (mm)",
        "Density (g/cm³)", "Lay Factor", "Weight kg/km"
    ]
    for col, h in enumerate(bom_headers, start=2):
        c = ws.cell(15, col, h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.border = THIN_BORDER
        c.alignment = CENTER_ALIGN

    # Pre-populate template rows (user fills in actual values)
    layer_templates = [
        (1, "Conductor",         "conductor",           "conductor_weight",   "=E8",     "—",  "=E8",     f"=VLOOKUP(E5&\"_conductor\",Master_Data!A:D,3,0)", "1.005", "=E6"),
        (2, "XLPE Insulation",   "xlpe_insulation",     "annular_layer",      "=P16",    "1.9","=Q17+2*R17","=VLOOKUP(C17,Master_Data!A:D,3,0)","1.000","=(PI()/4)*(S17^2-R17^2)*T17*U17"),
        (3, "Bedding",           "pvc_armoured_sheath", "annular_layer",      "=S17",    "0.9","=R18+2*R18","=VLOOKUP(C18,Master_Data!A:D,3,0)","1.008","=(PI()/4)*(S18^2-R18^2)*T18*U18"),
        (4, "GS Flat Strip Arm.","gs_flat_strip_armour","flat_strip_armour",  "=S18",    "0.8","=R19+2*R19","=VLOOKUP(C19,Master_Data!A:D,3,0)","1.008",""),
        (5, "FRLSH Outer Sheath","frlsh_outer_sheath",  "annular_layer",      "=S19",    "1.8","=R20+2*R20","=VLOOKUP(C20,Master_Data!A:D,3,0)","1.008","=(PI()/4)*(S20^2-R20^2)*T20*U20"),
    ]
    for i, (no, name, mat, formula, id_mm, thick, od, dens, lay, weight) in enumerate(layer_templates):
        row_no = 16 + i
        row_vals = [no, name, mat, formula, id_mm, thick, od, dens, lay, weight]
        for col, v in enumerate(row_vals, start=2):
            c = ws.cell(row_no, col, v)
            c.font = BODY_FONT
            c.border = THIN_BORDER
        if i % 2 == 1:
            for col in range(2, 12):
                ws.cell(row_no, col).fill = ALT_FILL

    total_row = 22
    ws.cell(total_row, 2).value = "TOTAL WEIGHT (kg/km)"
    ws.cell(total_row, 2).font = Font(name="Calibri", bold=True, size=10)
    ws.cell(total_row, 11).value = "=SUM(K16:K21)"
    ws.cell(total_row, 11).font = Font(name="Calibri", bold=True, size=10, color=RAVIN_BLUE)

    note = ws.cell(total_row + 2, 2)
    note.value = "Note: Replace formula placeholders with actual cell references after entering your cable specs. Add more rows as needed."
    note.font = Font(name="Calibri", italic=True, size=9, color="666666")

    _set_col_widths(ws, [4, 8, 24, 22, 18, 14, 14, 14, 16, 13, 13])


def build_costing_calculator(wb):
    ws = wb.create_sheet("Costing_Calculator")
    ws.sheet_properties.tabColor = "CC3300"

    ws["A1"] = "CABLE COSTING CALCULATOR"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="CC3300")

    ws["B3"] = "Link to BOM_Calculator sheet (BOM_Calculator!K16:K21) or paste weights below."
    ws["B3"].font = Font(name="Calibri", italic=True, size=10, color="666666")

    cost_headers = ["Layer Name", "Material Key", "Weight kg/km", "RM Price (₹/kg)", "Material Cost (₹/km)"]
    for col, h in enumerate(cost_headers, start=2):
        c = ws.cell(5, col, h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = THIN_BORDER; c.alignment = CENTER_ALIGN

    cost_templates = [
        ("Conductor",          "copper_conductor",    "=BOM_Calculator!K16", f"=VLOOKUP(C6,RM_Prices!A:C,3,0)",  "=D6*E6"),
        ("XLPE Insulation",    "xlpe_insulation",     "=BOM_Calculator!K17", f"=VLOOKUP(C7,RM_Prices!A:C,3,0)",  "=D7*E7"),
        ("Bedding",            "pvc_armoured_sheath", "=BOM_Calculator!K18", f"=VLOOKUP(C8,RM_Prices!A:C,3,0)",  "=D8*E8"),
        ("GS Flat Strip Arm.", "gs_flat_strip_armour","=BOM_Calculator!K19", f"=VLOOKUP(C9,RM_Prices!A:C,3,0)",  "=D9*E9"),
        ("FRLSH Outer Sheath", "frlsh_outer_sheath",  "=BOM_Calculator!K20", f"=VLOOKUP(C10,RM_Prices!A:C,3,0)", "=D10*E10"),
    ]
    for i, (name, mat, wt, price, cost) in enumerate(cost_templates):
        row_no = 6 + i
        for col, v in enumerate([name, mat, wt, price, cost], start=2):
            c = ws.cell(row_no, col, v); c.font = BODY_FONT; c.border = THIN_BORDER
        if i % 2 == 1:
            for col in range(2, 7):
                ws.cell(row_no, col).fill = ALT_FILL

    summary_row = 13
    ws.cell(summary_row, 2).value = "Total Material Cost (₹/km):"
    ws.cell(summary_row, 2).font = Font(name="Calibri", bold=True, size=10)
    ws.cell(summary_row, 6).value = "=SUM(F6:F11)"
    ws.cell(summary_row, 6).font = Font(name="Calibri", bold=True, size=11, color=RAVIN_BLUE)

    ws.cell(summary_row + 1, 2).value = "Drum Cost (₹/km):"
    ws.cell(summary_row + 1, 2).font = Font(name="Calibri", bold=True, size=10)
    ws.cell(summary_row + 1, 6).value = "0"
    ws.cell(summary_row + 1, 6).font = BODY_FONT

    ws.cell(summary_row + 2, 2).value = "Conversion Cost (₹/km):"
    ws.cell(summary_row + 2, 2).font = Font(name="Calibri", bold=True, size=10)
    ws.cell(summary_row + 2, 6).value = "0"
    ws.cell(summary_row + 2, 6).font = BODY_FONT

    ws.cell(summary_row + 3, 2).value = "Total Floor Cost (₹/km):"
    ws.cell(summary_row + 3, 2).font = Font(name="Calibri", bold=True, size=11, color="CC3300")
    ws.cell(summary_row + 3, 6).value = f"=F{summary_row}+F{summary_row+1}+F{summary_row+2}"
    ws.cell(summary_row + 3, 6).font = Font(name="Calibri", bold=True, size=11, color="CC3300")

    ws.cell(summary_row + 5, 2).value = "Margin (%):"
    ws.cell(summary_row + 5, 2).font = Font(name="Calibri", bold=True, size=10)
    ws.cell(summary_row + 5, 6).value = 15
    ws.cell(summary_row + 5, 6).font = BODY_FONT

    ws.cell(summary_row + 6, 2).value = "Selling Price (₹/km):"
    ws.cell(summary_row + 6, 2).font = Font(name="Calibri", bold=True, size=12, color="1F7A1F")
    ws.cell(summary_row + 6, 6).value = f"=F{summary_row+3}/(1-F{summary_row+5}/100)"
    ws.cell(summary_row + 6, 6).font = Font(name="Calibri", bold=True, size=12, color="1F7A1F")

    _set_col_widths(ws, [4, 24, 22, 14, 18, 20])


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    print("Building sheets...")
    build_readme(wb)
    print("  ✓ README")
    build_layer_registry(wb)
    print("  ✓ Layer_Registry")
    build_cable_families(wb)
    print("  ✓ Cable_Families")
    build_master_data(wb)
    print("  ✓ Master_Data")
    build_lay_factors(wb)
    print("  ✓ Lay_Factors")
    build_extrusion_tolerances(wb)
    print("  ✓ Extrusion_Tolerances")
    build_rm_prices(wb)
    print("  ✓ RM_Prices")
    build_drum_costs(wb)
    print("  ✓ Drum_Costs")
    build_margins(wb)
    print("  ✓ Margins")
    build_bom_calculator(wb)
    print("  ✓ BOM_Calculator")
    build_costing_calculator(wb)
    print("  ✓ Costing_Calculator")

    wb.save(OUTPUT_PATH)
    print(f"\nSaved: {OUTPUT_PATH}")
    print(f"Sheets: {', '.join(ws.title for ws in wb.worksheets)}")


if __name__ == "__main__":
    main()
