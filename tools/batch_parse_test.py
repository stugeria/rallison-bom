"""
Batch GTP Parser Test — processes a folder of GTP PDFs and outputs a review workbook.

Useful for validating parser output across 40-50 real GTPs before using in production.

Usage:
    python tools/batch_parse_test.py /path/to/gtp/folder [--gtp-type A]
    python tools/batch_parse_test.py /path/to/gtps --output results.xlsx

Output: parse_test_results.xlsx (or --output path) with 3 sheets:
  - Summary:       1 row per cable — GTP ref, item, designation, config, cable_type, R_DC, layers found
  - Layer_Details: 1 row per layer — all parsed thickness/material values for visual review
  - Errors:        any PDF that failed or returned no cables
"""

import argparse
import os
import sys
import traceback
from datetime import datetime

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BASE_DIR)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl>=3.1.0")
    sys.exit(1)

from core.gtp_parser_direct import parse_gtp_direct

# ── Style helpers (reuse from build_excel.py pattern) ───────────────────────
RAVIN_BLUE  = "003366"
ORANGE      = "E65C00"
GREEN       = "1F7A1F"
RED_BG      = "FDECEA"

def _hdr(ws, row, ncols, color=RAVIN_BLUE):
    fill = PatternFill("solid", fgColor=color)
    font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    for col in range(1, ncols + 1):
        c = ws.cell(row, col)
        c.font = font; c.fill = fill; c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")

def _data(ws, row, ncols, alt=False, highlight=False):
    fill = PatternFill("solid", fgColor=("FFF3E0" if highlight else ("F5F5F5" if alt else "FFFFFF")))
    font = Font(name="Calibri", size=10)
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin",  color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    for col in range(1, ncols + 1):
        c = ws.cell(row, col)
        c.font = font; c.fill = fill; c.border = border

def _widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Parser runner ────────────────────────────────────────────────────────────

def parse_folder(folder: str, gtp_type_override=None):
    pdfs = sorted(f for f in os.listdir(folder) if f.lower().endswith(".pdf"))
    if not pdfs:
        print(f"No PDFs found in: {folder}")
        sys.exit(1)

    print(f"Found {len(pdfs)} PDF(s) in {folder}\n")

    results = []   # list of (filename, result_dict, error_str)
    for i, filename in enumerate(pdfs, 1):
        path = os.path.join(folder, filename)
        print(f"  [{i:2d}/{len(pdfs)}] Parsing: {filename}", end="  ")
        try:
            result = parse_gtp_direct(path, gtp_type_override)
            n_cables = len(result.get("cables", []))
            print(f"→ {n_cables} cable(s) | type={result.get('gtp_type','?')} | ref={result.get('gtp_ref','?')}")
            results.append((filename, result, None))
        except Exception as e:
            print(f"→ ERROR: {e}")
            results.append((filename, None, traceback.format_exc()))

    return results


# ── Excel writer ─────────────────────────────────────────────────────────────

def write_results(results, output_path: str):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Summary ────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    ws_sum.freeze_panes = "A2"
    sum_hdrs = [
        "File", "GTP Ref", "Customer", "GTP Type",
        "Item No.", "Designation", "Config",
        "Conductor", "Cable Type", "Voltage kV",
        "Standard", "Cores", "R_DC (Ω/km)", "Num Wires",
        "Conductor OD (mm)", "Overall OD (mm)",
        "Delivery (m)", "Drum Type",
        "Layers Found", "Layer Keys", "Parse OK?"
    ]
    ws_sum.append(sum_hdrs)
    _hdr(ws_sum, 1, len(sum_hdrs))

    # ── Sheet 2: Layer Details ───────────────────────────────────────────────
    ws_lyr = wb.create_sheet("Layer_Details")
    ws_lyr.freeze_panes = "A2"
    lyr_hdrs = [
        "File", "GTP Ref", "Item No.", "Designation", "Config", "Cable Type",
        "Layer No.", "Layer Name", "Material Key", "Formula Type",
        "Nominal Thickness (mm)", "Thickness Type",
        "Strip Width (mm)", "Strip Thickness (mm)",
        "Wire Diameter (mm)", "Tape Thickness (mm)", "Tape Overlap (%)",
        "N Wires / N Pairs", "OD (mm)", "Notes"
    ]
    ws_lyr.append(lyr_hdrs)
    _hdr(ws_lyr, 1, len(lyr_hdrs), color=ORANGE)

    # ── Sheet 3: Errors ──────────────────────────────────────────────────────
    ws_err = wb.create_sheet("Errors")
    ws_err.freeze_panes = "A2"
    err_hdrs = ["File", "Error Type", "Error Detail"]
    ws_err.append(err_hdrs)
    _hdr(ws_err, 1, len(err_hdrs), color="AA0000")

    # Load layer registry for formula_type lookup
    import json
    _reg_path = os.path.join(BASE_DIR, "data", "layer_registry.json")
    with open(_reg_path) as f:
        layer_registry = json.load(f)

    sum_row = 2
    lyr_row = 2
    err_row = 2
    total_cables = 0
    total_layers = 0
    total_errors = 0

    for filename, result, error in results:
        if error:
            ws_err.cell(err_row, 1, filename)
            ws_err.cell(err_row, 2, error.split("\n")[-2] if "\n" in error else "ParseError")
            ws_err.cell(err_row, 3, error[:500])
            _data(ws_err, err_row, 3, highlight=True)
            err_row += 1
            total_errors += 1

            # Also add a summary row marking as failed
            ws_sum.cell(sum_row, 1, filename)
            for col in range(2, len(sum_hdrs)):
                ws_sum.cell(sum_row, col, "—")
            ws_sum.cell(sum_row, len(sum_hdrs), "NO — parse error")
            _data(ws_sum, sum_row, len(sum_hdrs), highlight=True)
            sum_row += 1
            continue

        cables = result.get("cables", [])
        if not cables:
            ws_err.cell(err_row, 1, filename)
            ws_err.cell(err_row, 2, "NoCablesFound")
            ws_err.cell(err_row, 3, "Parser ran OK but found no cable sections with valid R_DC")
            _data(ws_err, err_row, 3, highlight=True)
            err_row += 1
            total_errors += 1

        for cable in cables:
            total_cables += 1
            layers = cable.get("layers", [])
            layer_keys = [l.get("material_key", "?") for l in layers]

            # Summary row
            row_vals = [
                filename,
                result.get("gtp_ref", ""),
                result.get("customer", ""),
                result.get("gtp_type", ""),
                cable.get("item_no", ""),
                cable.get("designation", ""),
                cable.get("config", ""),
                cable.get("conductor_material", ""),
                cable.get("cable_type", ""),
                cable.get("voltage_kv", ""),
                cable.get("standard", ""),
                cable.get("num_cores", ""),
                cable.get("dc_resistance_ohm_per_km", ""),
                cable.get("num_wires", ""),
                cable.get("conductor_od_mm", ""),
                cable.get("overall_od_mm", ""),
                cable.get("delivery_length_m", ""),
                cable.get("drum_type", ""),
                len(layers),
                " | ".join(layer_keys),
                "YES" if cable.get("dc_resistance_ohm_per_km") else "WARN — no R_DC",
            ]
            alt = (total_cables % 2 == 0)
            warn = not cable.get("dc_resistance_ohm_per_km")
            for col, v in enumerate(row_vals, 1):
                ws_sum.cell(sum_row, col, v)
            _data(ws_sum, sum_row, len(sum_hdrs), alt=alt, highlight=warn)
            sum_row += 1

            # Layer detail rows
            for lno, layer in enumerate(layers, 1):
                total_layers += 1
                mat_key = layer.get("material_key", "")
                formula_type = layer_registry.get(mat_key, {}).get("formula_type", "")
                lyr_vals = [
                    filename,
                    result.get("gtp_ref", ""),
                    cable.get("item_no", ""),
                    cable.get("designation", ""),
                    cable.get("config", ""),
                    cable.get("cable_type", ""),
                    lno,
                    layer.get("layer_name", ""),
                    mat_key,
                    formula_type,
                    layer.get("nominal_thickness_mm", ""),
                    layer.get("thickness_type", ""),
                    layer.get("armour_strip_width_mm", ""),
                    layer.get("armour_strip_thickness_mm", ""),
                    layer.get("wire_diameter_mm", ""),
                    layer.get("tape_thickness_mm", ""),
                    layer.get("tape_overlap_pct", ""),
                    layer.get("n_wires") or layer.get("n_pairs", ""),
                    layer.get("od_mm", ""),
                    "missing thickness" if (
                        formula_type in ("annular_layer",) and
                        not layer.get("nominal_thickness_mm")
                    ) else "",
                ]
                alt_l = (lno % 2 == 0)
                warn_l = formula_type in ("annular_layer",) and not layer.get("nominal_thickness_mm")
                for col, v in enumerate(lyr_vals, 1):
                    ws_lyr.cell(lyr_row, col, v)
                _data(ws_lyr, lyr_row, len(lyr_hdrs), alt=alt_l, highlight=warn_l)
                lyr_row += 1

    # ── Column widths ────────────────────────────────────────────────────────
    _widths(ws_sum, [22, 16, 18, 8, 7, 18, 16, 12, 14, 10, 12, 7, 14, 10, 14, 14, 12, 10, 7, 50, 16])
    _widths(ws_lyr, [22, 14, 6, 16, 14, 12, 6, 22, 22, 18, 16, 14, 12, 14, 14, 14, 12, 12, 10, 20])
    _widths(ws_err, [22, 20, 80])

    # ── Stats sheet ──────────────────────────────────────────────────────────
    ws_stat = wb.create_sheet("Stats", 0)
    ws_stat["A1"] = "BATCH PARSE TEST — RESULTS"
    ws_stat["A1"].font = Font(name="Calibri", bold=True, size=14, color=RAVIN_BLUE)
    ws_stat["A2"] = f"Run: {datetime.now().strftime('%Y-%m-%d %H:%M')}   |   Folder: {os.path.abspath(args.folder)}"
    ws_stat["A2"].font = Font(name="Calibri", italic=True, size=10)

    stats = [
        ("PDFs processed",       len(results)),
        ("Total cables parsed",  total_cables),
        ("Total layers found",   total_layers),
        ("PDFs with errors",     total_errors),
        ("Avg layers / cable",   round(total_layers / max(total_cables, 1), 1)),
    ]
    for i, (label, val) in enumerate(stats, start=4):
        ws_stat.cell(i, 1, label).font = Font(name="Calibri", bold=True, size=11)
        ws_stat.cell(i, 2, val).font = Font(name="Calibri", size=11, color=RAVIN_BLUE)

    ws_stat["A10"] = "How to review:"
    ws_stat["A10"].font = Font(name="Calibri", bold=True, size=11)
    tips = [
        "1. Go to Summary sheet — scan 'Parse OK?' column. Yellow = warning (missing R_DC). Red = parse error.",
        "2. Check 'Cable Type' column — verify LT/HT/control/instrumentation was detected correctly.",
        "3. Check 'Layer Keys' column — confirm the right layers were found for each cable family.",
        "4. Go to Layer_Details — look at 'Nominal Thickness (mm)' — yellow = thickness was missing.",
        "5. For armoured cables: verify Strip Width/Thickness or Wire Diameter columns are populated.",
        "6. Go to Errors sheet — investigate any PDFs that failed completely.",
        "7. Check 'Num Wires' and 'Conductor OD' — if blank, the parser used fallback calculations.",
    ]
    for i, tip in enumerate(tips, start=11):
        ws_stat.cell(i, 1, tip).font = Font(name="Calibri", size=10)

    ws_stat.column_dimensions["A"].width = 85
    ws_stat.column_dimensions["B"].width = 20

    wb.save(output_path)
    print(f"\n{'='*60}")
    print(f"Parsed {len(results)} PDF(s) → {total_cables} cables, {total_layers} layers")
    print(f"Errors: {total_errors}")
    print(f"Output: {output_path}")
    print(f"{'='*60}")


# ── Main ─────────────────────────────────────────────────────────────────────

parser = argparse.ArgumentParser(description="Batch GTP parse test")
parser.add_argument("folder", help="Folder containing GTP PDFs")
parser.add_argument("--gtp-type", default=None, help="Force GTP type A/B/C (default: auto from filename)")
parser.add_argument("--output", default=None, help="Output Excel path (default: <folder>/parse_test_results.xlsx)")
args = parser.parse_args()

if not os.path.isdir(args.folder):
    print(f"ERROR: Not a directory: {args.folder}")
    sys.exit(1)

output_path = args.output or os.path.join(args.folder, "parse_test_results.xlsx")

results = parse_folder(args.folder, args.gtp_type)
write_results(results, output_path)
